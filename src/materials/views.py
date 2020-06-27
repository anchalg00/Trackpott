from django.shortcuts import render,redirect,render_to_response
from .forms import MaterialRegisterForm,ProjectRegisterForm,DocumentRegisterForm,StoreRegisterForm,ProgressRegisterForm
from django.views.generic import ListView,CreateView,TemplateView
from django.core import serializers
from .models import Requirement,Proj,Subproj,Document,Store,Progress,Schedule
from django.http import JsonResponse, HttpResponse
from django.forms import modelformset_factory
from django.contrib.auth.decorators import login_required
from datetime import date
from fpdf import FPDF
from django.urls import reverse
from django.http import HttpResponseRedirect
from io import BytesIO
from django.core.files import File
from django.contrib import messages
from datetime import datetime
from openpyxl import Workbook
from helpers.Render import Render
# from .utils import render_to_pdf
# from .forms import UploadFileForm

# from chartit import DataPool, Chart, PivotDataPool, PivotChart
# Create your views here.
@login_required()
def register_spools_view(request,*args,**kwargs):

	project_selected = Schedule.objects.filter(is_selected=True)
	selected_project = project_selected.first()
	object_list = Requirement.objects.filter(proj = project_selected.first())

	if request.method=='POST':
			form = MaterialRegisterForm(request.POST)
			if form.is_valid():
				obj= form.save()				
				obj.proj = project_selected.first()
				obj.save()
				username=form.cleaned_data.get('username');
				
				return redirect('RequirementList')
	else:
		form=MaterialRegisterForm()
	return render(request,"materials/listview.html",{'form':form ,'object_list' : object_list ,'project_list_first': selected_project})

@login_required()
def register_progress_view(request,*args,**kwargs):
	project_selected = Schedule.objects.filter(is_selected=True)
	selected_project = project_selected.first()

	object_list = Progress.objects.filter(proj = project_selected.first())

	if request.method=='POST':
			form = ProgressRegisterForm(request.POST)
			if form.is_valid():
				obj= form.save()				
				obj.proj = project_selected.first()
				obj.save()
				username=form.cleaned_data.get('username');
				
				return redirect('Progress')
	else:
		form=ProgressRegisterForm()
	return render(request,"materials/listviewProgress.html",{'form':form ,'object_list' : object_list,'project_list_first': selected_project})

@login_required()
def register_store_view(request,*args,**kwargs):
	project_selected = Schedule.objects.filter(is_selected=True)
	selected_project = project_selected.first()

	object_list = Store.objects.filter()

	if request.method=='POST':
			form1 = StoreRegisterForm(request.POST, request.FILES)
			if form1.is_valid():
				obj= form1.save()
				obj.save()				
				
				pdf = FPDF()
				pdf.add_page()
				pdf.rect(5, 5, 200, 287, 'D')
				pdf.set_font("Times", style='B', size=25)
				pdf.image('static/logo.png', x=5, y=5, w=40)
				pdf.cell(200, 30, txt='MATERIAL TRANSFER', ln=1, align="C")
				pdf.image('static/logo.png', x=165, y=5, w=40)
				pdf.set_font("Arial", size=15)
				order = str(obj.PO_no_s)
				Dt = str(date.today())
				data = [[str(obj.item_s), str(obj.spec_s), str(obj.material_s), str(obj.qty_s) ]]
				pdf.cell(200, 20, txt='', ln=1, align="L")
				pdf.cell(200, 10, txt='Order No:   ' + order, ln=1, align="L")
				pdf.cell( 200,10, txt='Date:   '+ Dt, ln=1, align="L")
				pdf.cell(200, 30, txt='', ln=1, align="L")
				pdf.set_line_width(0.4)
				col_width = pdf.w / 4.5
				row_height = pdf.font_size
				data1=[['Item','Specifications','Material','Quantity']]
				for row in data1:
				    for item in row:
				        pdf.cell(col_width*1.0, row_height*1.5, txt=item, align="C", border=1)
				    pdf.ln(row_height*1.5)
				for row in data:
				    for item in row:
				        pdf.cell(col_width*1.0, row_height*1.5, txt=item, align="C", border=1)
				    pdf.ln(row_height*1.5)
				pdfname=order+'.pdf'
				pdfname2='../media/'+pdfname
				pdf.output(pdfname2)
				obj.pdf_s=pdfname
				obj.save()
				inst=obj.pdf_s
				messages.success(request,inst,'Report generated')

				return HttpResponseRedirect(reverse('Store',args=()))

	else:
		form1=StoreRegisterForm()
	return render(request,"materials/listviewStore.html",{'form1':form1 ,'object_list' : object_list,'project_list_first': selected_project})

@login_required()
def documents_view(request,*args,**kwargs):
	project_selected = Schedule.objects.filter(is_selected=True)
	selected_project = project_selected.first()
	object_list = Document.objects.filter(proj = project_selected.first())
	form = DocumentRegisterForm(request.POST,request.FILES)

	if request.method=='POST':
		if request.POST['_id'] == '0':
			if form.is_valid():
				form.instance=form.save(commit=False)
				form.instance.save()
				return redirect('/documentview/')
		else:
			initial_id=request.POST['_id']
			print(initial_id)
			document=Document.objects.get(pk=initial_id)
			form = DocumentRegisterForm(request.POST,request.FILES,instance=document)
			if form.is_valid():
				form.save()
				return redirect('/documentview/')

	

	if 'ADNOC_documentno' in request.GET and request.GET['ADNOC_documentno']!='':
		object_list=Document.objects.filter(ADNOC_documentno=request.GET['ADNOC_documentno'],proj=project_selected.first())

	if form.is_valid():
			obj= form.save()
			
			obj.proj = project_selected.first()
			obj.save()
			username=form.cleaned_data.get('username');
			
			return redirect('/documentview/')
	else:
		form=DocumentRegisterForm()
		

	return render(request,"materials/listviewDocsss.html",{'form':form ,'object_list' : object_list,'project_list_first': selected_project})


@login_required
def get_doc(request, id):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    project_item = Document.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", project_item), safe=False)




# class table_view(ListView):
# 	template_name="materials/listview.html"
# 	queryset = Requirement.objects.all()

# 	if request.method=='POST':
# 			form = MaterialRegisterForm(request.POST)
# 			if form.is_valid():
# 				form.save()
# 				username=form.cleaned_data.get('username');
				
# 				return redirect('RequirementList')
# 	else:
# 		form=MaterialRegisterForm()
# 	return render(request,"materials/listview.html",{'form':form })


@login_required()
def subproject_view(request,page_id):
	obj = Proj.objects.get(id=page_id)
	
	SubprojectFormset = modelformset_factory(Subproj , fields = ('name','plan_date','complete_date'))
	#ACTIVITES FORMSET
	if request.method == 'POST':
			# formset = SubprojectFormset(request.POST , queryset=Subproj.objects.filter(proj_id =obj.id , is_subproj = False))
			formset = SubprojectFormset(request.POST , queryset=Subproj.objects.filter(proj_id =obj.id))
			if formset.is_valid():
				instances = formset.save(commit = False)
				for instance in instances:
					instance.proj_id = obj.id
					# instance.is_subproj = False
					instance.save()
				return redirect('pagedetail',page_id=obj.id)
	# formset = SubprojectFormset(queryset=Subproj.objects.filter(proj_id=obj.id , is_subproj=False))
	formset = SubprojectFormset(queryset=Subproj.objects.filter(proj_id=obj.id ))
	#SUBPROJECT FORMSET

	# if request.method == 'POST':
	# 	formset2 = SubprojectFormset(request.POST , queryset=Subproj.objects.filter(proj_id =obj.id , is_subproj=True))
	# 	if formset2.is_valid():
	# 		instances2 = formset2.save(commit = False)
	# 		for instance2 in instances2:
	# 			instance2.proj_id = obj.id
	# 			instance2.is_subproj = True
	# 			instance2.save()
	# 		return redirect('pagedetail',page_id=obj.id)
	# formset2 = SubprojectFormset(queryset=Subproj.objects.filter(proj_id=obj.id, is_subproj=True))
	
	# sales = DataPool(
	#    series=
	#     [{'options': {
	#     #    'source': SalesReport.objects.all()},
	#     'source': Subproj.objects.filter(proj_id = obj.id)},
	#     #'source': SalesReport.objects.filter(sales__lte=10.00)},
	#         'terms': [{'name': 'name',
	#         'days': 'days'}]
	#         },
	#     ]) 





	# cht = Chart(
	#    	datasource = sales,
	#     series_options =
	#       [{'options':{
	#           'type': 'column',
	#           'stacking': False},
	#         'terms':{
	#           'name': [
	#             'days']
	#         }}],
	#     chart_options =
	#       {'title': {
	#            'text': 'Activity Days'},
	#        'xAxis': {
	#            'title':{'text': 'Activity name'}},
	#        'yAxis': {
	#            'title': {'text': 'Number of days'}},
	#        })
	
              
	# qs = Subproj.objects.filter(proj_id = obj.id)
	
	# import plotly.figure_factory as ff
	# from plotly.offline import plot

	# df = [dict(Task="Job A", Start='2009-01-01', Finish='2009-02-28'),
	#       dict(Task="Job B", Start='2009-03-05', Finish='2009-04-15'),
	#       dict(Task="Job C", Start='2009-02-20', Finish='2009-05-30')]

	# fig = ff.create_gantt(df)
	# figure = plot(fig)

	context = {
		"object" : obj ,
		'formset': formset ,
		# 'chart_list' : cht,
		# 'qs':qs,
		# 'graph':figure,
		# 'formset2': formset2
		}


	return render_to_response('materials/create_subproject.html', context) 
	# return render(request,"materials/create_subproject.html",context)

@login_required
def excel_export_doc(request):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    queryset_list = Document.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-document-report-trackpott.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Report - Document'

    # Define the titles for columns
    columns = [
      	'ADNOC_documentno',
	    'description',
	    'revision',
	    'due_date',
	    'submission_date',
	    'transmittal_number',
	    'reason_of_issue',
	    'document',
	    'aprroval_date',
	    'receipt_transmittalno',
	    # 'view_response',
	    'aprroval_code',
	    'remarks']
    
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for qs in queryset_list:
        row_num += 1

        # Define the data for each cell in the row
        # http://127.0.0.1:8000/media/documents/2020/05/19/2K18-CO-065_Assignment2_uAmmiU5.pdf
        row = [
            qs.ADNOC_documentno,
            qs.description,
            qs.revision,
            str(qs.due_date),
            str(qs.submission_date),
            qs.transmittal_number,
            qs.reason_of_issue,
            str("http://127.0.0.1:8000/media/"+ str(qs.document)),
            str(qs.aprroval_date),
            qs.receipt_transmittalno,
            # qs.view_response,
            qs.aprroval_code,
            qs.remarks
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response





