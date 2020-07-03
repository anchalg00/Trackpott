from django.shortcuts import render, redirect
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from datetime import datetime
from openpyxl import Workbook
from helpers.Render import Render
from .models import Store
from .forms import MaterialForm
from django.contrib.auth.decorators import login_required


@login_required
def index(request):
    initial_id = 0
    form = MaterialForm(request.POST or None)
    if request.method == 'POST':
        if request.POST['_id'] == '0':
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.created_by_id = request.user.id
                form_instance.save()
                return redirect('materials:index')
        else:
            initial_id = request.POST['_id']
            store = Store.objects.get(pk=initial_id)
            form = MaterialForm(request.POST or None, instance=store)
            if form.is_valid():
                form.save()
                return redirect('materials:index')
    query = """
    SELECT
        *,
        IFNULL((SELECT SUM(pi.quantity) AS po_quantity FROM purchase_orders_purchaseorderitem AS pi JOIN purchase_orders_purchaseorder AS p ON p.id = pi.purchase_order_id WHERE p.po_status = 'd' AND pi.item_id = ms.id),0) AS po_quantity,
        IFNULL((SELECT SUM(quantity) AS p_quantity FROM projects_projectitem WHERE is_approved = 1 AND item_id = ms.id),0) AS p_quantity
    FROM
        materials_store AS ms WHERE id != 0 
        """
    if 'item_s' in request.GET and request.GET['item_s'] != '':
        query += " AND ms.item_s LIKE %s" % "'%%"+request.GET['item_s']+"%%'"
    if 'material_s' in request.GET and request.GET['material_s'] != '':
        query += " AND ms.material_s LIKE %s" % "'%%" + \
            request.GET['material_s']+"%%'"
    query += ' ORDER BY ms.item_s ASC'
    materials = Store.objects.raw(str(query))
    context = {
        "materials": materials,
        "form": form,
        "initial_id": initial_id
    }
    return render(request, 'materials/manage_materials.html', context)


@login_required
def get_material(request, id):
    store = Store.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", store), safe=False)


@login_required
def excel_export(request):
    queryset_list = Store.objects.raw("""
    SELECT
        *,
        IFNULL((SELECT SUM(pi.quantity) AS po_quantity FROM purchase_orders_purchaseorderitem AS pi JOIN purchase_orders_purchaseorder AS p ON p.id = pi.purchase_order_id WHERE p.po_status = 'd' AND pi.item_id = ms.id),0) AS po_quantity,
        IFNULL((SELECT SUM(quantity) AS p_quantity FROM projects_projectitem WHERE is_approved = 1 AND item_id = ms.id),0) AS p_quantity
    FROM
        materials_store AS ms ORDER BY ms.item_s ASC
        """)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-material-report-trackpott.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Report - Materials'

    # Define the titles for columns
    columns = [
        'Item',
        'Specification',
        'Material',
        'Rating',
        'Size 1',
        'Schedule 1',
        'Size 2',
        'Schedule 2',
        'Facing',
        'Quantity'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for qs in queryset_list:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            qs.item_s,
            qs.spec_s,
            qs.material_s,
            qs.rating_s,
            qs.size1_s,
            qs.sch1_s,
            qs.size2_s,
            qs.sch2_s,
            qs.facing_s,
            int(qs.po_quantity) - int(qs.p_quantity)
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
