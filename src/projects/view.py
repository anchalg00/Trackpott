from django.shortcuts import render, redirect
from django.core import serializers
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponse
from chart.models import Schedule
from django.contrib.auth.models import User
from .models import Project, ProjectItem, ProjectMissingItem, ProjectActivity, ProjectActivityPlan, ProjectActivityPlanLog, ScheduleInfo, ScheduleDocument
from .forms import ProjectForm, ProjectItemForm, ProjectActivityForm, ProjectActivityFormSet, ProjectActivityPlanForm, ScheduleDocumentForm, ScheduleInfoForm
from materials.models import Store
from datetime import datetime
from openpyxl import Workbook
from helpers.Render import Render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from purchase_orders.models import PurchaseOrderItem
from django.db import connection as conn


@login_required
def index(request):
    id = Schedule.objects.filter(is_selected=True).first().id
    if id is not None:
        initial_id = 0
        form = ProjectForm(request.POST or None, request=request)
        if request.method == 'POST':
            if request.POST['_id'] == '0':
                if form.is_valid():
                    form_instance = form.save(commit=False)
                    form_instance.schedule_id = id
                    # form_instance.created_by_id = request.user.id
                    form_instance.save()
                    return redirect('projects:index')
            else:
                initial_id = request.POST['_id']
                project = Project.objects.get(pk=initial_id)
                form = ProjectForm(request.POST or None,
                                   instance=project, request=request)
                if form.is_valid():
                    form_instance = form.save(commit=False)
                    form_instance.save()
                    return redirect('projects:index')
        queryset_list = Project.objects.filter(
            schedule_id=id).order_by('receipt_number')
        if 'receipt_number' in request.GET and request.GET['receipt_number'] != '':
            queryset_list = queryset_list.filter(
                receipt_number__icontains=request.GET['receipt_number'])
        if 'name' in request.GET and request.GET['name'] != '':
            queryset_list = queryset_list.filter(
                schedule__name__icontains=request.GET['name'])
        context = {
            "projects": queryset_list,
            "schedule": Schedule.objects.get(id=id),
            "form": form,
            "initial_id": initial_id
        }
        return render(request, 'projects/manage_projects.html', context)
    else:
        return HttpResponse('<h1>Project is not selected</h1>')


@login_required
def get_project(request, id):
    project = Project.objects.filter(id=id)
    return JsonResponse({"project": list(project.values()), "username": User.objects.filter(id=project[0].assigned_for_id).first().username})


@login_required
def transfer(request):
    id = Schedule.objects.filter(is_selected=True).first().id
    if id is not None:
        initial_id = 0
        form = ProjectItemForm(request.POST or None)
        if request.method == 'POST':
            if request.POST['_id'] == '0':
                if form.is_valid():
                    form_instance = form.save(commit=False)
                    form_instance.project_id = id
                    form_instance.created_by_id = request.user.id
                    form_instance.save()
                    return redirect('projects:items', id=id)
            else:
                initial_id = request.POST['_id']
                project_item = ProjectItem.objects.get(pk=initial_id)
                form = ProjectItemForm(
                    request.POST or None, instance=project_item)
                if form.is_valid():
                    form.save()
                    return redirect('projects:items', id=id)
        context = {
            "project_items": ProjectItem.objects.filter(project_id=id).order_by('-id'),
            "form": form,
            "initial_id": initial_id,
            "project": Project.objects.get(id=id)
        }
        return render(request, 'projects/manage_project_items.html', context)
    else:
        return HttpResponse('<h1>Project is not selected</h1>')


@login_required
def items(request, id):
    initial_id = 0
    form = ProjectItemForm(request.POST or None)
    if request.method == 'POST':
        if request.POST['_id'] == '0':
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.project_id = id
                form_instance.created_by_id = request.user.id
                form_instance.save()
                return redirect('projects:items', id=id)
        else:
            initial_id = request.POST['_id']
            project_item = ProjectItem.objects.get(pk=initial_id)
            form = ProjectItemForm(request.POST or None, instance=project_item)
            if form.is_valid():
                form.save()
                return redirect('projects:items', id=id)
    project = Project.objects.get(id=id)
    context = {
        "project_items": ProjectItem.objects.filter(project_id=id).order_by('-id'),
        "form": form,
        "initial_id": initial_id,
        "schedule": Schedule.objects.get(id=project.schedule_id),
        "project": project
    }
    return render(request, 'projects/manage_project_items.html', context)


@login_required
def get_project_item(request, id):
    project_item = ProjectItem.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", project_item), safe=False)


@login_required
def transfer_report(request):
    queryset_list = ProjectItem.objects.order_by('-id')
    if 'item_id' in request.GET and request.GET['item_id'] != '':
        queryset_list = queryset_list.filter(item_id=request.GET['item_id'])
    if 'from_date' in request.GET and request.GET['from_date'] != '':
        from_date = datetime.strptime(request.GET['from_date'], '%d-%m-%Y')
        if from_date:
            queryset_list = queryset_list.filter(created_at__gte=from_date)
    if 'to_date' in request.GET and request.GET['to_date'] != '':
        to_date = datetime.strptime(request.GET['to_date'], '%d-%m-%Y')
        if to_date:
            queryset_list = queryset_list.filter(created_at__lte=to_date)
    context = {
        'materials': Store.objects.all(),
        'items': queryset_list
    }
    if 'export' in request.GET and request.GET['export'] == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-material-transfer-report-trackpott.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Report - Material Transfer'

        # Define the titles for columns
        columns = [
            'Date of transfer',
            'Item Name',
            'Specification',
            'Material',
            'Quantity',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for qs in queryset_list:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                qs.created_at,
                qs.item.item_s,
                qs.item.spec_s,
                qs.item.material_s,
                qs.quantity
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    elif 'export' in request.GET and request.GET['export'] == 'pdf':
        return Render.render('projects/pdf_material_transfer_report.html', context)
    else:
        return render(request, 'projects/material_transfer_report.html', context)


@login_required
def pdf_report(request, id):
    try:
        project = Project.objects.get(pk=id)
    except Project.DoesNotExist:
        return HttpResponse('<h1>Page not found!</h1>')

    project_items = ProjectItem.objects.filter(project_id=id)
    context = {
        'project': project,
        'project_items': project_items
    }
    return Render.render('projects/pdf_report.html', context)


@login_required
def pending_material_report(request):
    queryset_list = ProjectItem.objects.filter(is_pending=True)
    if 'receipt_number' in request.GET and request.GET['receipt_number'] != '':
        queryset_list = queryset_list.filter(
            project__receipt_number=request.GET['receipt_number'])
    context = {
        "pending_items": queryset_list
    }
    if 'export' in request.GET and request.GET['export'] == 'pdf':
        return Render.render('projects/pdf_pending_material_report.html', context)
    else:
        return render(request, 'projects/pending_material_report.html', context)


@login_required
def missing_material_report(request):
    queryset_list = ProjectMissingItem.objects.all()
    if 'receipt_number' in request.GET and request.GET['receipt_number'] != '':
        queryset_list = queryset_list.filter(
            project__receipt_number=request.GET['receipt_number'])
    context = {
        "missing_items": queryset_list
    }
    if 'export' in request.GET and request.GET['export'] == 'pdf':
        return Render.render('projects/pdf_missing_material_report.html', context)
    else:
        return render(request, 'projects/missing_material_report.html', context)


@login_required
def approval_request(request):
    queryset_list = ProjectItem.objects.filter(
        project__assigned_for_id=request.user.id, is_approved=False)
    context = {
        "pending_items": queryset_list
    }
    return render(request, 'projects/approval_request.html', context)


@login_required
def approve_item(request, id):
    try:
        _quantity = request.GET.get('quantity')
        project_item = ProjectItem.objects.get(pk=id)
        qs_po = PurchaseOrderItem.objects.values('item_id').filter(purchase_order__po_status='d').filter(
            item_id=project_item.item_id).annotate(sum_quantity=Sum('quantity'))
        qs_project = ProjectItem.objects.values('item_id').filter(is_approved=True).filter(
            item_id=project_item.item_id).annotate(sum_quantity=Sum('quantity'))
        po_quantity = qs_po[0]['sum_quantity'] if qs_po is not None and qs_po else 0
        project_quantity = qs_project[0]['sum_quantity'] if qs_project is not None and qs_project else 0
        actual_quantity = int(po_quantity) - int(project_quantity)
        if project_item.quantity < int(_quantity):
            return render(request, 'projects/stock_not_available.html')
        if actual_quantity < int(_quantity):
            return render(request, 'projects/stock_not_available.html')
        else:
            if project_item.quantity > int(_quantity):
                project_quantity = int(project_item.quantity) - int(_quantity)
                ProjectMissingItem.objects.create(
                    quantity=project_quantity,
                    project_id=project_item.project_id,
                    item_id=project_item.item_id,
                    created_by_id=request.user.id
                )
            project_item.quantity = int(_quantity)
            project_item.is_pending = False
            project_item.is_approved = True
            project_item.save()
            return redirect('projects:approval_request')
    except ProjectItem.DoesNotExist:
        return HttpResponse('<h1>Item not found!</h1>')


@login_required
def project_activity(request):
    id = Schedule.objects.filter(is_selected=True).first().id
    if ProjectActivity.objects.filter(schedule_id=id).count() != 0:
        if id is not None:
            initial_id = 0
            form = ProjectActivityForm(request.POST or None)
            if request.method == 'POST':
                if request.POST['_id'] == '0':
                    if form.is_valid():
                        form_instance = form.save(commit=False)
                        # form_instance.weightage = float(
                        #     request.POST['weightage'])
                        form_instance.schedule_id = id
                        # form_instance.revisor_id = int(request.POST['revisor_id'])
                        # form_instance.created_by_id = request.user.id
                        form_instance.save()
                        return redirect('projects:project_activity')
                else:
                    initial_id = request.POST['_id']
                    project_activity = ProjectActivity.objects.get(
                        pk=initial_id)
                    form = ProjectActivityForm(request.POST or None,
                                               instance=project_activity)
                    if form.is_valid():
                        form_instance = form.save(commit=False)
                        form_instance.save()
                        return redirect('projects:project_activity')
            queryset_list = ProjectActivity.objects.filter(schedule_id=id)
            if 'name' in request.GET and request.GET['name'] != '':
                queryset_list = queryset_list.filter(
                    name=request.GET['name'])
            context = {
                "project_activities": queryset_list,
                "form": form,
                "initial_id": initial_id,
                "schedule": Schedule.objects.get(id=id),
                "users": User.objects.all(),
                "total_weightage": ProjectActivity.objects.filter(schedule_id=id).aggregate(Sum('weightage')),
                "revisor": ProjectActivity.objects.filter(schedule_id=id).first()
            }
            return render(request, 'projects/activity/manage_activity.html', context)
        else:
            return HttpResponse('<h1>Project is not selected</h1>')
    return redirect('projects:create_activity')


@login_required
def get_project_activity(request, id):
    project_activity = ProjectActivity.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", project_activity), safe=False)


@login_required
def create_activity(request):
    id = Schedule.objects.filter(is_selected=True).first().id
    if ProjectActivity.objects.filter(schedule_id=id).count() == 0:
        if id is not None:
            InitialProjectActivityFormSet = modelformset_factory(
                ProjectActivity, form=ProjectActivityForm, extra=1)
            formset = InitialProjectActivityFormSet(
                request.POST or None, queryset=ProjectActivity.objects.none())
            if request.method == 'POST':
                formset = modelformset_factory(
                    ProjectActivity, form=ProjectActivityForm, formset=ProjectActivityFormSet, extra=int(request.POST['form-TOTAL_FORMS']))(request.POST)
                # formset = ProjectActivityFormSet(request.POST)
                if formset.is_valid():
                    for form in formset:
                        ProjectActivity.objects.create(
                            name=form.cleaned_data.get('name'),
                            weightage=form.cleaned_data.get('weightage'),
                            schedule_id=id
                        )
                    return redirect('projects:create_activity')
                else:
                    context = {
                        "schedule": Schedule.objects.get(id=id),
                        "formset": formset
                    }
                    return render(request, 'projects/activity/create_activity.html', context)
            context = {
                "schedule": Schedule.objects.get(id=id),
                "formset": formset
            }
            return render(request, 'projects/activity/create_activity.html', context)
        else:
            return HttpResponse('<h1>Project is not selected</h1>')
    return redirect('projects:project_activity')


@login_required
def edit_activity(request):
    id = Schedule.objects.filter(is_selected=True).first().id
    total_activities = ProjectActivity.objects.filter(schedule_id=id).count()
    if total_activities > 0:
        if id is not None:
            InitialProjectActivityFormSet = modelformset_factory(
                ProjectActivity, can_delete=True, form=ProjectActivityForm, extra=0)
            formset = InitialProjectActivityFormSet(
                request.POST or None, queryset=ProjectActivity.objects.filter(schedule_id=id))
            if request.method == 'POST':
                formset = modelformset_factory(
                    ProjectActivity, can_delete=True, form=ProjectActivityForm, formset=ProjectActivityFormSet, extra=int(request.POST['form-TOTAL_FORMS']))(request.POST)
                # # formset.errors
                print(formset.errors)
                if formset.is_valid():
                    # print('d')
                    new_items = []
                    # for form in formset:
                    #     print(form.cleaned_data.get('id').id)
                    #     npa = ProjectActivity.objects.filter(
                    #         id=form.cleaned_data.get('id').id, schedule_id=id).first()
                    #     if npa is not None:
                    #         if npa.id not in new_items:
                    #             new_items.append(npa.id)
                    # print("new_items", new_items)
                    for form in formset:
                        # pass
                        if form.cleaned_data.get('id') is not None:
                            ProjectActivity.objects.filter(id=form.cleaned_data.get('id').id).update(
                                name=form.cleaned_data.get('name'),
                                weightage=form.cleaned_data.get('weightage')
                            )
                        else:
                            ProjectActivity.objects.create(
                                name=form.cleaned_data.get('name'),
                                weightage=form.cleaned_data.get('weightage'),
                                schedule_id=id
                            )
                    return redirect('projects:create_activity')
                else:
                    # print(formset.errors)
                    context = {
                        "schedule": Schedule.objects.get(id=id),
                        "formset": formset
                    }
                    return render(request, 'projects/activity/edit_activity.html', context)
            context = {
                "schedule": Schedule.objects.get(id=id),
                "formset": formset
            }
            return render(request, 'projects/activity/edit_activity.html', context)
        else:
            return HttpResponse('<h1>Project is not selected</h1>')
    return redirect('projects:create_activity')


@login_required
def plans(request, id):
    initial_id = 0
    project_activity = ProjectActivity.objects.get(id=id)
    schedule = Schedule.objects.get(id=project_activity.schedule_id)
    start_year = datetime.strptime(str(schedule.start_date), "%Y-%m-%d").year
    end_year = start_year
    if schedule.end_date is not None:
        end_year = datetime.strptime(str(schedule.end_date), "%Y-%m-%d").year
    _end_year = end_year if start_year != end_year else int(end_year)+1
    plan_year_options = tuple([(r, r)
                               for r in range(start_year, _end_year)])
    form = ProjectActivityPlanForm(
        request.POST or None, request=request, project_activity_id=id, plan_year_options=plan_year_options, schedule_start_date=schedule.start_date)
    if request.method == 'POST':
        if request.POST['_id'] == '0':
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.project_activity_id = id
                form_instance.created_by_id = request.user.id
                form_instance.save()
                return redirect('projects:plans', id=id)
        else:
            initial_id = request.POST['_id']
            project_activity_plan = ProjectActivityPlan.objects.get(
                pk=initial_id)
            form = ProjectActivityPlanForm(
                request.POST or None, request=request, project_activity_id=id, instance=project_activity_plan)
            if form.is_valid():
                form.save()
                return redirect('projects:plans', id=id)
    context = {
        "project_activity_plans": ProjectActivityPlan.objects.filter(project_activity_id=id).order_by('-id'),
        "form": form,
        "initial_id": initial_id,
        "schedule": schedule,
        "project_activity": project_activity
    }
    return render(request, 'projects/plan/manage_plans.html', context)


@login_required
def get_project_activity_plan(request, id):
    project_activity_plan = ProjectActivityPlan.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", project_activity_plan), safe=False)


@login_required
def set_plan_fixed(request, id):
    try:
        project_activity_plan = ProjectActivityPlan.objects.get(id=id)
        project_activity_plan.is_plan_fixed = True
        project_activity_plan.save()
        return redirect('projects:plans', id=project_activity_plan.project_activity_id)
    except ProjectActivityPlan.DoesNotExist:
        return HttpResponse('Page not found')


@login_required
def update_revisor(request, id):
    user_id = request.GET['user_id']
    ProjectActivity.objects.filter(
        schedule_id=id).update(revisor_id=user_id)
    return redirect('projects:project_activity')


@login_required
def revise_plans(request):
    queryset_list = ProjectActivityPlan.objects.filter(
        project_activity__revisor_id=request.user.id)

    if 'project' in request.GET and request.GET['project'] != '':
        queryset_list = queryset_list.filter(
            project_activity__schedule_id=request.GET['project'])
    if 'activity_name' in request.GET and request.GET['activity_name'] != '':
        queryset_list = queryset_list.filter(
            project_activity__name__icontains=request.GET['activity_name'])
    context = {
        "schedules": Schedule.objects.all(),
        "project_activity_plans": queryset_list
    }
    return render(request, 'projects/plan/revise_plans.html', context)


@login_required
def revise_plan(request, id):
    try:
        _plan_value = request.GET.get('plan_value')
        project_activity_plan = ProjectActivityPlan.objects.get(pk=id)
        if project_activity_plan.is_plan_fixed == True:
            ProjectActivityPlanLog.objects.create(
                old_plan_value=project_activity_plan.plan_value,
                project_activity_plan_id=project_activity_plan.id,
                created_by_id=request.user.id

            )
        project_activity_plan.plan_value = _plan_value
        project_activity_plan.save()
        return redirect('projects:revise_plans')
    except ProjectActivityPlan.DoesNotExist:
        return HttpResponse('<h1>Plan not found!</h1>')

    context = {
        "projecT_activity_plans": queryset_list
    }
    return render(request, 'projects/plan/revise_plans.html', context)


@login_required
def revised_plan_report(request):
    queryset_list = ProjectActivityPlanLog.objects.all()
    if 'project' in request.GET and request.GET['project'] != '':
        queryset_list = queryset_list.filter(
            project_activity_plan__project_activity__schedule_id=request.GET['project'])
    if 'activity_name' in request.GET and request.GET['activity_name'] != '':
        queryset_list = queryset_list.filter(
            project_activity_plan__project_activity__name__icontains=request.GET['activity_name'])
    context = {
        "schedules": Schedule.objects.all(),
        "project_activity_plan_logs": queryset_list
    }
    return render(request, 'projects/plan/revised_plan_report.html', context)


months = ["Unknown", "Jan", "Feb", "Mar", "Apr", "May",
          "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def month_year_iter(start_month, start_year, end_month, end_year):
    ym_start = 12*start_year + start_month - 1
    ym_end = 12*end_year + end_month - 1
    for ym in range(ym_start, ym_end):
        y, m = divmod(ym, 12)
        yield y, months[m+1]


def excel_activity(request):
    schedule = Schedule.objects.filter(is_selected=True).first()
    if schedule.id:
        if ProjectActivity.objects.filter(schedule_id=schedule.id).count() > 0:
            st = datetime.strptime(str(schedule.start_date), '%Y-%m-%d')
            if schedule.end_date is not None:
                ed = datetime.strptime(str(schedule.end_date), '%Y-%m-%d')
                ed_month = ed.month
                ed_year = ed.year
                if ed.year == st.year and ed.month == st.month:
                    ed_month = ed.month + 1
            else:
                ed_month = int(st.month)+12
                ed_year = st.year
            list_years = list(month_year_iter(
                st.month, st.year, ed_month, ed_year))
            project_activities = ProjectActivity.objects.filter(
                schedule_id=schedule.id)
            listing = []
            i = 0
            for project_activity in project_activities:
                listing.append(
                    {"project_activity": project_activity, "plans": []})
                project_activity_plans = ProjectActivityPlan.objects.filter(
                    project_activity_id=project_activity.id).order_by('plan_year', 'plan_month')
                j = 0
                for project_activity_plan in project_activity_plans:
                    listing[i]["plans"].append(project_activity_plan)
                    j = j+1
                i = i+1
            context = {
                "start_year": st.year,
                "end_year": ed_year,
                "start_month": st.month,
                "end_month": ed_month,
                "activities": listing,
                "schedule": schedule,
                "list_years": list_years
            }
            return render(request, 'projects/activity/activity_excel_view.html', context)
        else:
            return redirect('projects:create_activity')
    return HttpResponse('<h1>Project is not selected</h1>')


@login_required
def get_project_activity_plan_by_year_month(request):
    project_activity_plan = ProjectActivityPlan.objects.filter(
        project_activity_id=request.GET['id'], plan_year=request.GET['plan_year'], plan_month=request.GET['plan_month'])
    return JsonResponse(serializers.serialize("json", project_activity_plan), safe=False)


@login_required
def update_project_activity_plan(request):
    project_activity_plan = ProjectActivityPlan.objects.filter(
        project_activity_id=request.GET['plan_activity_id'], plan_year=request.GET['plan_year'], plan_month=request.GET['plan_month']).first()
    if project_activity_plan:
        if project_activity_plan.is_plan_fixed == True:
            ProjectActivityPlanLog.objects.create(
                old_plan_value=project_activity_plan.plan_value,
                project_activity_plan_id=project_activity_plan.id,
                created_by_id=request.user.id

            )
        project_activity_plan.plan_value = request.GET['plan_value']
        project_activity_plan.actual_value = request.GET[
            'actual_value'] if request.GET['actual_value'] != '' else 0.00
        project_activity_plan.save()
    else:
        ProjectActivityPlan.objects.create(
            project_activity_id=request.GET['plan_activity_id'],
            plan_year=request.GET['plan_year'],
            plan_month=request.GET['plan_month'],
            plan_value=request.GET['plan_value'],
            actual_value=request.GET['actual_value']
        )
    project_activity_plan = ProjectActivityPlan.objects.filter(
        project_activity_id=request.GET['plan_activity_id'], plan_year=request.GET['plan_year'], plan_month=request.GET['plan_month'])
    return JsonResponse(serializers.serialize("json", project_activity_plan), safe=False)


def chart_report(request):
    schedule = Schedule.objects.filter(is_selected=True).first()
    if schedule.id:
        form = ScheduleDocumentForm(
            request.POST or None, request.FILES or None)
        if request.method == 'POST':
            if form.is_valid():
                schedule_document_file = form.save(commit=False)
                schedule_document_file.schedule_id = schedule.id
                schedule_document_file.save()
                return redirect('projects:chart_report')
        query = """
        SELECT ppa.id, ppap.id, ppap.plan_month, ppap.plan_year, SUM(ppap.plan_value) AS sum_plan_value, SUM(ppap.actual_value) AS sum_actual_value FROM projects_projectactivity ppa JOIN projects_projectactivityplan ppap ON ppa.id = ppap.project_activity_id WHERE ppa.schedule_id = '"""+str(schedule.id)+"""' GROUP BY ppap.plan_year, ppap.plan_month ORDER BY ppap.plan_year, SUBSTRING(str_to_date(CONCAT('2012-',ppap.plan_month,'-01'),'%Y-%b-%d'),6,2) ASC
            """
        project_activities = Schedule.objects.raw(str(query))
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        year_months = []
        plan_value = []
        actual_value = []
        last_actual_value = 0.00
        for row in rows:
            year_months.append(str(row[3]+"-"+row[2]))
            plan_value.append(row[4])
            if row[5]:
                actual_value.append(row[5])
                last_actual_value = row[5]
            else:
                actual_value.append(last_actual_value)
            schedule_info = ScheduleInfo.objects.filter(
                schedule_id=schedule.id).first()
        context = {
            "schedule_info": schedule_info,
            "form": form,
            "schedule": schedule,
            "schedule_documents": ScheduleDocument.objects.filter(schedule_id=schedule.id),
            "year_months": year_months,
            "plan_value": plan_value,
            "actual_value": actual_value,
        }
        return render(request, 'projects/activity/activity_chart.html', context)
    return HttpResponse('<h1>Project is not selected</h1>')


@login_required
def edit_schedule_info(request):
    schedule = Schedule.objects.filter(is_selected=True).first()
    if schedule is not None:
        schedule_info = ScheduleInfo.objects.filter(
            schedule_id=schedule.id).first()
        if schedule_info is None:
            form = ScheduleInfoForm(request.POST or None)
        else:
            form = ScheduleInfoForm(
                request.POST or None, instance=schedule_info)
        if request.method == 'POST':
            if form.is_valid():
                if schedule_info is not None:
                    form.save()
                else:
                    form_instance = form.save(commit=False)
                    form_instance.schedule_id = schedule.id
                    form_instance.save()
                return redirect('projects:chart_report')
        context = {
            "form": form,
            "schedule_info": schedule_info,
            "schedule": schedule
        }
        return render(request, 'projects/activity/edit_schedule_info.html', context)
    return HttpResponse('<h1>Project is not selected</h1>')
