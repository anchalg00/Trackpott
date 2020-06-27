from django.shortcuts import render, redirect
from django.core import serializers
from django.http import JsonResponse, HttpResponse
from .models import PurchaseOrder, PurchaseOrderItem
from .forms import PurchaseOrderForm, PurchaseOrderItemForm
from chart.models import Schedule
from materials.models import Store
from datetime import datetime
from openpyxl import Workbook
from helpers.Render import Render
from django.contrib.auth.decorators import login_required


@login_required
def index(request):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    initial_id = 0
    form = PurchaseOrderForm(request.POST or None)
    if request.method == 'POST':
        if request.POST['_id'] == '0':
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.created_by_id = request.user.id
                form_instance.save()
                return redirect('purchase_orders:index')
        else:
            initial_id = request.POST['_id']
            purchase_order = PurchaseOrder.objects.get(pk=initial_id)
            form = PurchaseOrderForm(
                request.POST or None, instance=purchase_order)
            if form.is_valid():
                form.save()
                return redirect('purchase_orders:index')

    queryset_list = PurchaseOrder.objects.all().order_by('po_number')
    if 'po_number' in request.GET and request.GET['po_number'] != '':
        queryset_list = queryset_list.filter(
            po_number__icontains=request.GET['po_number'])
    context = {
        "purchase_orders": queryset_list,
        "form": form,
        "initial_id": initial_id,
        'project_selected' : project_selected,
        'project_list_first': selected_project
    }
    return render(request, 'purchase_orders/manage_purchase_orders.html', context)


@login_required
def get_purchase_order(request, id):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    purchase_order = PurchaseOrder.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", purchase_order), safe=False)


@login_required
def items(request, id):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    initial_id = 0
    form = PurchaseOrderItemForm(request.POST or None)
    if request.method == 'POST':
        if request.POST['_id'] == '0':
            if form.is_valid():
                form_instance = form.save(commit=False)
                form_instance.purchase_order_id = id
                form_instance.created_by_id = request.user.id
                form_instance.save()
                return redirect('purchase_orders:items', id=id)
        else:
            initial_id = request.POST['_id']
            purchase_order_item = PurchaseOrderItem.objects.get(pk=initial_id)
            form = PurchaseOrderItemForm(
                request.POST or None, instance=purchase_order_item)
            if form.is_valid():
                form.save()
                return redirect('purchase_orders:items', id=id)
    context = {
        "purchase_order_items": PurchaseOrderItem.objects.filter(purchase_order_id=id).order_by('-id'),
        "form": form,
        "initial_id": initial_id,
        "purchase_order": PurchaseOrder.objects.get(id=id),
        'project_selected' : project_selected,
        'project_list_first': selected_project
    }
    return render(request, 'purchase_orders/manage_purchase_order_items.html', context)


@login_required
def get_purchase_order_item(request, id):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    purchase_order_item = PurchaseOrderItem.objects.filter(id=id)
    return JsonResponse(serializers.serialize("json", purchase_order_item), safe=False)


@login_required
def report(request):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    queryset_list = PurchaseOrder.objects.order_by('-id')
    if 'delivery_date_from' in request.GET and request.GET['delivery_date_from'] != '':
        delivery_date_from = datetime.strptime(
            request.GET['delivery_date_from'], '%d-%m-%Y')
        if delivery_date_from:
            queryset_list = queryset_list.filter(
                delivery_date__gte=delivery_date_from)
    if 'delivery_date_to' in request.GET and request.GET['delivery_date_to'] != '':
        delivery_date_to = datetime.strptime(
            request.GET['delivery_date_to'], '%d-%m-%Y')
        if delivery_date_to:
            queryset_list = queryset_list.filter(
                delivery_date__lte=delivery_date_to)
    context = {
        'purchase_orders': queryset_list,
        'project_selected' : project_selected,
        'project_list_first': selected_project
    }

    if 'export' in request.GET and request.GET['export'] == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-purchase-order-report-trackpott.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Report - Purchase Orders'

        # Define the titles for columns
        columns = [
            'PO No.',
            'Delivery Date',
            'Status'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for qs in queryset_list:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                qs.po_number,
                qs.delivery_date,
                qs.po_status
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    elif 'export' in request.GET and request.GET['export'] == 'pdf':
        return Render.render('purchase_orders/pdf_po_report.html', context)
    else:
        return render(request, 'purchase_orders/report.html', context)


@login_required
def addition_report(request):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    queryset_list = PurchaseOrderItem.objects.order_by('-id')
    if 'item_id' in request.GET and request.GET['item_id'] != '':
        queryset_list = queryset_list.filter(item_id=request.GET['item_id'])
    if 'from_date' in request.GET and request.GET['from_date'] != '':
        from_date = datetime.strptime(request.GET['from_date'], '%d-%m-%Y')
        if from_date:
            queryset_list = queryset_list.filter(created_at__gte=from_date)
    if 'to_date' in request.GET and request.GET['to_date'] != '':
        to_date = datetime.strptime(request.GET['to_date'], '%d-%m-%Y')
        if to_date:
            queryset_list = queryset_list.filter(created_at__lte=to_date)
    context = {
        'materials': Store.objects.all(),
        'items': queryset_list,
        'project_selected' : project_selected,
        'project_list_first': selected_project
    }
    if 'export' in request.GET and request.GET['export'] == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-material-addition-report-trackpott.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Report - Material Addition'

        # Define the titles for columns
        columns = [
            'Date of addition',
            'Item Name',
            'Specification',
            'Material',
            'Quantity',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for qs in queryset_list:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                qs.created_at,
                qs.item.item_s,
                qs.item.spec_s,
                qs.item.material_s,
                qs.quantity
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    elif 'export' in request.GET and request.GET['export'] == 'pdf':
        return Render.render('purchase_orders/pdf_material_addition_report.html', context)
    else:
        return render(request, 'purchase_orders/material_addition_report.html', context)


@login_required
def pdf_report(request, id):
    project_selected = Schedule.objects.filter(is_selected=True)
    selected_project = project_selected.first()
    try:
        purchase_order = PurchaseOrder.objects.get(pk=id)
    except PurchaseOrder.DoesNotExist:
        return HttpResponse('<h1>Page not found!</h1>')

    purchase_order_items = PurchaseOrderItem.objects.filter(
        purchase_order_id=id)
    context = {
        'purchase_order': purchase_order,
        'purchase_order_items': purchase_order_items,
        'project_selected' : project_selected,
        'project_list_first': selected_project
    }
    return Render.render('purchase_orders/pdf_report.html', context)
